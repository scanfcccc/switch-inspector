import csv
import io
from typing import List, Dict, Optional
from engine.field_catalog import FieldDef


def export_csv(
    rows: List[Dict],
    fields: List[FieldDef],
    selected_keys: List[str],
) -> bytes:
    output = io.StringIO()
    field_defs = {f.key: f for f in fields}

    # determine columns
    cols = [k for k in selected_keys if k in field_defs] if selected_keys else [k for k in rows[0].keys() if k in field_defs]
    # also include device meta
    meta_cols = ['_device_ip', '_device_name', '_device_serial']
    for mc in meta_cols:
        if mc not in cols and any(mc in r for r in rows):
            cols.insert(0, mc)

    labels = []
    for c in cols:
        fd = field_defs.get(c)
        labels.append(fd.label if fd else c)

    writer = csv.writer(output)
    writer.writerow(labels)

    for row in rows:
        writer.writerow([row.get(c, '') for c in cols])

    return output.getvalue().encode('utf-8-sig')


def export_xlsx(
    rows: List[Dict],
    fields: List[FieldDef],
    selected_keys: List[str],
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "巡检数据"

    field_defs = {f.key: f for f in fields}
    cols = [k for k in selected_keys if k in field_defs] if selected_keys else [k for k in rows[0].keys() if k in field_defs]
    meta_cols = ['_device_ip', '_device_name', '_device_serial']
    for mc in meta_cols:
        if mc not in cols and any(mc in r for r in rows):
            cols.insert(0, mc)

    labels = []
    for c in cols:
        fd = field_defs.get(c)
        labels.append(fd.label if fd else c)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for ci, label in enumerate(labels, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = header_font
        cell.fill = header_fill

    for ri, row in enumerate(rows, 2):
        for ci, col in enumerate(cols, 1):
            ws.cell(row=ri, column=ci, value=row.get(col, ''))

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
